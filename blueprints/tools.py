# -*- coding: utf-8 -*-
"""Blueprint для инструментов"""
import io
import xlwt
import xlrd
from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required, current_user
from datetime import datetime
from openpyxl import load_workbook
from io import BytesIO
from models import PurchasePrice, db
from utils.constants import MOSCOW_TZ
from utils.margin import load_user_margin_settings, save_user_margin_settings
from utils.api import fetch_warehouses_data

tools_bp = Blueprint('tools', __name__)


@tools_bp.route("/tools/labels", methods=["GET"]) 
@login_required
def tools_labels_page():
    """Страница генерации этикеток"""
    token = (current_user.wb_token or "") if current_user.is_authenticated else ""
    error = None
    if not token:
        error = "Укажите токен API в профиле"
    return render_template("tools_labels.html", error=error, token=token)


@tools_bp.route("/tools/prices", methods=["GET"])
@login_required
def tools_prices_page():
    """Страница управления ценами"""
    from utils.margin import load_user_margin_settings
    from utils.constants import MOSCOW_TZ
    from datetime import datetime
    
    token = current_user.wb_token or ""
    error = None
    products = []
    prices_data = {}
    commission_data = {}
    dimensions_data = {}
    warehouses_data = []
    stocks_data = {}
    fbs_stocks_data = {}
    purchase_prices: dict[str, float] = {}
    
    # Загружаем настройки маржи
    margin_settings = load_user_margin_settings(current_user.id)
    
    if not token:
        error = "Укажите токен API в профиле"
    else:
        try:
            # Загружаем товары из кэша или с API
            from utils.cache import load_products_cache, save_products_cache
            from utils.api import fetch_all_cards
            from utils.helpers import normalize_cards_response
            cached = load_products_cache()
            if cached and cached.get("_user_id") == current_user.id:
                products = cached.get("items", [])
            else:
                # Загружаем все страницы товаров
                raw_cards = fetch_all_cards(token, page_limit=100)
                products = normalize_cards_response({"cards": raw_cards})
                save_products_cache({"items": products, "_user_id": current_user.id})
            
            # Получаем цены продажи для товаров
            if products:
                nm_ids = []
                for p in products:
                    nm_id = p.get("nm_id")
                    if nm_id:
                        try:
                            nm_ids.append(int(nm_id))
                        except (ValueError, TypeError):
                            continue
                
                if nm_ids:
                    from utils.api import fetch_prices_data
                    prices_data = fetch_prices_data(token, nm_ids)
            
            # Получаем данные о комиссиях
            try:
                from utils.api import fetch_commission_data
                commission_data = fetch_commission_data(token)
                print(f"Загружено {len(commission_data)} комиссий")
            except Exception as e:
                print(f"Ошибка при загрузке комиссий: {e}")
                commission_data = {}
            
            # Получаем данные о размерах товаров из карточек
            try:
                for product in products:
                    nm_id = product.get('nm_id')
                    dimensions = product.get('dimensions', {})
                    if nm_id and dimensions:
                        dimensions_data[nm_id] = dimensions
                print(f"Загружено {len(dimensions_data)} записей размеров из карточек")
            except Exception as e:
                print(f"Ошибка при обработке размеров: {e}")
                dimensions_data = {}
            
            # Получаем данные о складах
            try:
                from utils.api import fetch_warehouses_data
                warehouses_data = fetch_warehouses_data(token)
                print(f"Загружено {len(warehouses_data)} складов")
            except Exception as e:
                print(f"Ошибка при загрузке складов: {e}")
                warehouses_data = []
            
            # Получаем данные об остатках FBW
            try:
                from utils.cache import load_stocks_cache
                stocks_cached = load_stocks_cache()
                if stocks_cached and stocks_cached.get("_user_id") == current_user.id:
                    items = stocks_cached.get("items", [])
                    for stock_item in items:
                        barcode = stock_item.get("barcode")
                        qty = int(stock_item.get("qty", 0) or 0)
                        if barcode:
                            if barcode not in stocks_data:
                                stocks_data[barcode] = 0
                            stocks_data[barcode] += qty
                    print(f"Загружено остатков для {len(stocks_data)} товаров")
            except Exception as e:
                print(f"Ошибка при загрузке остатков: {e}")
                stocks_data = {}

            # Получаем данные об остатках FBS
            try:
                from utils.api import fetch_fbs_warehouses, fetch_fbs_stocks_by_warehouse
                prod_cached = load_products_cache() or {}
                products_all = prod_cached.get("products") or prod_cached.get("items") or []
                skus = []
                for p in products_all:
                    if isinstance(p.get("barcodes"), list):
                        skus.extend([str(x) for x in p.get("barcodes") if x])
                    elif p.get("barcode"):
                        skus.append(str(p.get("barcode")))
                skus = list({s for s in skus if s})
                if skus:
                    wlist = fetch_fbs_warehouses(token)
                    for w in wlist or []:
                        wid = w.get("id") or w.get("warehouseId") or w.get("warehouseID")
                        if not wid:
                            continue
                        try:
                            stocks = fetch_fbs_stocks_by_warehouse(token, int(wid), skus)
                        except Exception:
                            stocks = []
                        for s in stocks or []:
                            bc = str(s.get("sku") or s.get("barcode") or "").strip()
                            amount = int(s.get("amount") or 0)
                            if bc:
                                fbs_stocks_data[bc] = fbs_stocks_data.get(bc, 0) + amount
                print(f"Загружено FBS остатков для {len(fbs_stocks_data)} товаров")
            except Exception as e:
                print(f"Ошибка при загрузке FBS остатков: {e}")
                fbs_stocks_data = {}

            # Получаем сохраненные закупочные цены из базы данных
            try:
                from models import PurchasePrice
                saved_prices = PurchasePrice.query.filter_by(user_id=current_user.id).all()
                print(f"Найдено {len(saved_prices)} записей закупочных цен в БД для пользователя {current_user.id}")
                for price_record in saved_prices:
                    # Баркод всегда храним и сравниваем как строку
                    if price_record.barcode:
                        barcode_str = str(price_record.barcode).strip()
                        # Убираем .0 в конце, если это целое число
                        if barcode_str.endswith('.0'):
                            try:
                                int(float(barcode_str))
                                barcode_str = barcode_str[:-2]
                            except (ValueError, TypeError):
                                pass
                        purchase_prices[barcode_str] = float(price_record.price)
                print(f"Загружено {len(purchase_prices)} сохраненных закупочных цен из БД")
                if purchase_prices:
                    # Логируем первые 5 баркодов для отладки
                    sample_barcodes = list(purchase_prices.keys())[:5]
                    print(f"Примеры загруженных баркодов из БД: {sample_barcodes}")
                # Также логируем примеры баркодов из товаров для сравнения
                if products:
                    product_barcodes = [str(p.get('barcode', '')).strip() for p in products[:5] if p.get('barcode')]
                    print(f"Примеры баркодов из товаров: {product_barcodes}")
            except Exception as e:
                print(f"Ошибка при загрузке закупочных цен: {e}")
                import traceback
                traceback.print_exc()
                purchase_prices = {}
        except Exception as exc:
            error = f"Ошибка: {exc}"
    
    # Время последнего обновления цен на момент рендера страницы
    prices_last_updated = datetime.now(MOSCOW_TZ).strftime("%d.%m.%Y %H:%M")
    
    return render_template(
        "tools_prices.html",
        error=error,
        products=products,
        prices_data=prices_data,
        commission_data=commission_data,
        dimensions_data=dimensions_data,
        warehouses_data=warehouses_data,
        stocks_data=stocks_data,
        fbs_stocks_data=fbs_stocks_data,
        purchase_prices=purchase_prices,
        margin_settings=margin_settings,
        prices_last_updated=prices_last_updated,
        token=token
    )


@tools_bp.route("/api/prices/upload", methods=["POST"])
@login_required
def api_prices_upload():
    """Загрузка закупочных цен из Excel файла"""
    try:
        if 'file' not in request.files:
            return jsonify({"success": False, "error": "Файл не найден"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"success": False, "error": "Файл не выбран"}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({"success": False, "error": "Поддерживаются только Excel файлы (.xlsx, .xls)"}), 400
        
        prices = {}
        updated_count = 0
        
        if file.filename.lower().endswith('.xlsx'):
            workbook = load_workbook(file, data_only=True)
            worksheet = workbook.active
            for row in worksheet.iter_rows(min_row=2, max_col=2, values_only=True):
                if len(row) >= 2 and row[0] and row[1]:
                    barcode = str(row[0]).strip()
                    # Убираем .0 в конце, если это целое число
                    if barcode.endswith('.0'):
                        try:
                            int(float(barcode))
                            barcode = barcode[:-2]
                        except (ValueError, TypeError):
                            pass
                    try:
                        price = float(row[1])
                        if price > 0:
                            prices[barcode] = price
                            updated_count += 1
                    except (ValueError, TypeError):
                        continue
        else:
            file.seek(0)
            workbook = xlrd.open_workbook(file_contents=file.read())
            worksheet = workbook.sheet_by_index(0)
            for row_idx in range(1, worksheet.nrows):
                if worksheet.ncols >= 2:
                    barcode_cell = worksheet.cell_value(row_idx, 0)
                    price_cell = worksheet.cell_value(row_idx, 1)
                    if barcode_cell and price_cell:
                        barcode = str(barcode_cell).strip()
                        # Убираем .0 в конце, если это целое число
                        if barcode.endswith('.0'):
                            try:
                                int(float(barcode))
                                barcode = barcode[:-2]
                            except (ValueError, TypeError):
                                pass
                        try:
                            price = float(price_cell)
                            if price > 0:
                                prices[barcode] = price
                                updated_count += 1
                        except (ValueError, TypeError):
                            continue
        
        return jsonify({
            "success": True,
            "prices": prices,
            "updated_count": updated_count
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": f"Ошибка обработки файла: {str(e)}"}), 500


@tools_bp.route("/api/prices/save", methods=["POST"])
@login_required
def api_prices_save():
    """Сохранение закупочных цен"""
    try:
        data = request.get_json()
        prices = data.get('prices', {})
        
        if not prices:
            return jsonify({"success": False, "error": "Нет данных для сохранения"}), 400
        
        saved_count = 0
        
        for barcode, price in prices.items():
            try:
                # Нормализуем баркод: всегда строка, убираем пробелы и .0 в конце
                barcode_str = str(barcode).strip()
                # Убираем .0 в конце, если это целое число (например, "123.0" -> "123")
                if barcode_str.endswith('.0'):
                    try:
                        # Проверяем, что это действительно целое число
                        int(float(barcode_str))
                        barcode_str = barcode_str[:-2]
                    except (ValueError, TypeError):
                        pass
                
                existing_price = PurchasePrice.query.filter_by(
                    user_id=current_user.id, 
                    barcode=barcode_str
                ).first()
                
                if existing_price:
                    existing_price.price = float(price)
                    existing_price.updated_at = datetime.now(MOSCOW_TZ)
                else:
                    new_price = PurchasePrice(
                        user_id=current_user.id,
                        barcode=barcode_str,
                        price=float(price)
                    )
                    db.session.add(new_price)
                
                saved_count += 1
                
            except Exception as e:
                print(f"Ошибка при сохранении цены для баркода {barcode}: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        db.session.commit()
        
        # Проверяем, что цены действительно сохранились
        verify_count = PurchasePrice.query.filter_by(user_id=current_user.id).count()
        print(f"После сохранения: в БД {verify_count} записей закупочных цен для пользователя {current_user.id}")
        
        return jsonify({
            "success": True,
            "saved_count": saved_count,
            "message": f"Сохранено {saved_count} цен"
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": f"Ошибка сохранения: {str(e)}"}), 500


@tools_bp.route("/tools/prices/template", methods=["GET"])
@login_required
def download_prices_template():
    """Скачать шаблон Excel для загрузки закупочных цен (XLS)."""
    try:
        workbook = xlwt.Workbook(encoding='utf-8')
        worksheet = workbook.add_sheet('Шаблон закупочных цен')

        header_style = xlwt.easyxf('font: bold on; align: horiz center;')
        text_style = xlwt.easyxf('align: horiz left;')

        worksheet.write(0, 0, 'Баркод', header_style)
        worksheet.write(0, 1, 'Закупочная цена', header_style)

        worksheet.write(1, 0, '2001234567890', text_style)
        worksheet.write(1, 1, '123.45', text_style)

        worksheet.col(0).width = 5000
        worksheet.col(1).width = 5000

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='Шаблон_закупочных_цен.xls',
            mimetype='application/vnd.ms-excel'
        )
    except Exception as e:
        return jsonify({"error": f"Не удалось создать шаблон: {str(e)}"}), 500


@tools_bp.route("/api/tools/prices/margin-settings", methods=["GET", "POST"])
@login_required
def api_margin_settings():
    """Получение/сохранение настроек маржи текущего пользователя."""
    try:
        if request.method == "GET":
            settings = load_user_margin_settings(current_user.id)
            return jsonify({"success": True, "settings": settings})
        else:
            data = request.get_json() or {}
            saved = save_user_margin_settings(current_user.id, data)
            return jsonify({"success": True, "settings": saved})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@tools_bp.route("/api/tools/prices/warehouses", methods=["GET"])
@login_required
def api_tools_prices_warehouses():
    """Возвращает список складов с коэффициентами для выпадающего списка."""
    try:
        token = current_user.wb_token or ""
        if not token:
            return jsonify({"success": False, "error": "Не указан WB токен"}), 400
        warehouses = fetch_warehouses_data(token)
        return jsonify({"success": True, "warehouses": warehouses})
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


