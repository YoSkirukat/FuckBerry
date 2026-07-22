# -*- coding: utf-8 -*-
"""Лента заказов — список заказов с детальной модалкой."""
import time
from datetime import datetime
from typing import Optional, Tuple

from flask import Blueprint, jsonify, render_template, request
from flask_login import current_user, login_required

from utils.cache import get_orders_with_period_cache
from utils.helpers import parse_date
from utils.order_feed import (
    STATUS_LABELS,
    ORDERS_LOOKUP_BEFORE_SALE_DAYS,
    SALES_LOOKUP_EXTRA_DAYS,
    build_feed_items,
    collect_orders_from_period_cache,
    collect_sales_from_period_cache,
    default_date_range,
    extend_iso_date,
    iso_date_in_range,
    load_finance_srid_index,
    update_finance_srid_index_from_api,
    update_sales_period_cache_from_api,
)
from utils.wb_token import effective_wb_api_token

order_feed_bp = Blueprint("order_feed", __name__)

# Короткий кэш готовой ленты для пакетной подгрузки offset/limit
_FEED_LIST_CACHE: dict[str, dict] = {}
_FEED_LIST_TTL_SEC = 120


def _parse_optional_range(from_key: str, to_key: str) -> Optional[Tuple[str, str]]:
    date_from = (request.args.get(from_key) or "").strip()
    date_to = (request.args.get(to_key) or "").strip()
    if not date_from or not date_to:
        return None
    try:
        parse_date(date_from)
        parse_date(date_to)
    except Exception:
        return None
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    return date_from, date_to


def _feed_cache_key(
    user_id: int,
    date_from: str,
    date_to: str,
    sale_from: str,
    sale_to: str,
    status: str,
    scheme: str,
    q: str,
) -> str:
    return f"{user_id}|{date_from}|{date_to}|{sale_from}|{sale_to}|{status}|{scheme}|{q}"


def _get_cached_feed(key: str):
    entry = _FEED_LIST_CACHE.get(key)
    if not entry:
        return None
    if (time.time() - entry.get("ts", 0)) > _FEED_LIST_TTL_SEC:
        _FEED_LIST_CACHE.pop(key, None)
        return None
    return entry


def _resolve_windows(
    order_range: Optional[Tuple[str, str]],
    sale_range: Optional[Tuple[str, str]],
) -> Tuple[str, str, str, str]:
    """
    Возвращает (orders_from, orders_to, sales_from, sales_to)
    для выборки из кэша / обновления API.
    """
    if order_range and sale_range:
        o_from, o_to = order_range
        s_from, s_to = sale_range
        sales_from = min(o_from, s_from)
        sales_to = max(extend_iso_date(o_to, SALES_LOOKUP_EXTRA_DAYS), s_to)
        return o_from, o_to, sales_from, sales_to
    if order_range:
        o_from, o_to = order_range
        return o_from, o_to, o_from, extend_iso_date(o_to, SALES_LOOKUP_EXTRA_DAYS)
    # только период выкупов
    s_from, s_to = sale_range  # type: ignore[misc]
    o_from = extend_iso_date(s_from, -ORDERS_LOOKUP_BEFORE_SALE_DAYS)
    return o_from, s_to, s_from, s_to


@order_feed_bp.route("/order-feed", methods=["GET"])
@login_required
def order_feed_page():
    return render_template(
        "order_feed.html",
        date_from="",
        date_to="",
        sale_from="",
        sale_to="",
        status_labels=STATUS_LABELS,
    )


def _apply_feed_filters(
    items: list,
    sale_range: Optional[Tuple[str, str]],
    status_filter: str,
    scheme_filter: str,
    q: str,
) -> list:
    if sale_range:
        sf, st = sale_range
        items = [
            it for it in items
            if iso_date_in_range(it.get("sold_at") or (it.get("sale") or {}).get("sold_at"), sf, st)
        ]
    if status_filter:
        items = [it for it in items if it.get("status") == status_filter]
    if scheme_filter in ("FBW", "FBS"):
        items = [it for it in items if it.get("scheme") == scheme_filter]
    if q:
        def _match(it: dict) -> bool:
            hay = " ".join(
                str(x or "")
                for x in (
                    it.get("gnumber"),
                    it.get("srid"),
                    it.get("article"),
                    it.get("name"),
                    it.get("barcode"),
                    it.get("nm_id"),
                    it.get("sticker"),
                )
            ).lower()
            return q in hay

        items = [it for it in items if _match(it)]
    return items


def _load_feed_list(
    *,
    refresh: bool = False,
) -> tuple[list, dict, bool, bool, bool, str, str, str, str]:
    """
    Собирает отфильтрованную ленту.
    Возвращает:
      items, cache_meta, has_sales, has_finance, refreshed,
      date_from, date_to, sale_from, sale_to
    """
    order_range = _parse_optional_range("date_from", "date_to")
    sale_range = _parse_optional_range("sale_from", "sale_to")
    if not order_range and not sale_range:
        raise ValueError("Укажите период заказов или период выкупов")

    status_filter = (request.args.get("status") or "").strip().lower()
    q = (request.args.get("q") or "").strip().lower()
    scheme_filter = (request.args.get("scheme") or "").strip().upper()

    date_from = order_range[0] if order_range else ""
    date_to = order_range[1] if order_range else ""
    sale_from = sale_range[0] if sale_range else ""
    sale_to = sale_range[1] if sale_range else ""

    orders_from, orders_to, sales_from, sales_to = _resolve_windows(order_range, sale_range)

    user_id = current_user.id
    token = effective_wb_api_token(current_user)
    cache_key = _feed_cache_key(
        user_id, date_from, date_to, sale_from, sale_to, status_filter, scheme_filter, q
    )

    if refresh:
        _FEED_LIST_CACHE.pop(cache_key, None)

    cached = None if refresh else _get_cached_feed(cache_key)
    if cached:
        return (
            cached["items"],
            cached.get("cache_meta") or {},
            cached.get("has_sales", False),
            cached.get("has_finance", False),
            False,
            date_from,
            date_to,
            sale_from,
            sale_to,
        )

    cache_meta: dict = {}
    refreshed = False

    if refresh and token:
        try:
            orders, cache_meta = get_orders_with_period_cache(
                token, orders_from, orders_to, bypass_today_ttl=True
            )
            refreshed = True
        except Exception as e:
            orders, cache_meta = collect_orders_from_period_cache(
                user_id, orders_from, orders_to
            )
            cache_meta = dict(cache_meta or {})
            cache_meta["refresh_error"] = str(e)
        try:
            update_sales_period_cache_from_api(user_id, token, sales_from, sales_to)
        except Exception as e:
            cache_meta = dict(cache_meta or {})
            cache_meta["sales_refresh_error"] = str(e)
        try:
            update_finance_srid_index_from_api(user_id, token, sales_from, sales_to)
        except Exception as e:
            cache_meta = dict(cache_meta or {})
            cache_meta["finance_refresh_error"] = str(e)
    else:
        orders, cache_meta = collect_orders_from_period_cache(
            user_id, orders_from, orders_to
        )

    sales = collect_sales_from_period_cache(user_id, sales_from, sales_to)
    finance_idx = load_finance_srid_index(user_id)
    items = build_feed_items(user_id, orders, sales, record_history=True)
    items = _apply_feed_filters(items, sale_range, status_filter, scheme_filter, q)

    has_sales = bool(sales)
    has_finance = bool(finance_idx)
    _FEED_LIST_CACHE[cache_key] = {
        "ts": time.time(),
        "items": items,
        "cache_meta": cache_meta,
        "has_sales": has_sales,
        "has_finance": has_finance,
    }
    return (
        items,
        cache_meta,
        has_sales,
        has_finance,
        refreshed,
        date_from,
        date_to,
        sale_from,
        sale_to,
    )


@order_feed_bp.route("/api/order-feed", methods=["GET"])
@login_required
def api_order_feed():
    """Список заказов за период из кэша статистики (+ продажи для статусов)."""
    refresh = request.args.get("refresh") in ("1", "true", "yes")
    limit = request.args.get("limit", type=int) or 100
    offset = request.args.get("offset", type=int) or 0
    limit = max(1, min(limit, 500))
    offset = max(0, offset)

    try:
        (
            items,
            cache_meta,
            has_sales,
            has_finance,
            refreshed,
            date_from,
            date_to,
            sale_from,
            sale_to,
        ) = _load_feed_list(refresh=refresh)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    total = len(items)
    page = items[offset : offset + limit]

    return jsonify({
        "items": page,
        "total": total,
        "offset": offset,
        "limit": limit,
        "date_from": date_from,
        "date_to": date_to,
        "sale_from": sale_from,
        "sale_to": sale_to,
        "refreshed": refreshed,
        "cache": cache_meta,
        "has_sales": has_sales,
        "has_finance": has_finance,
        "updated_at": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
    })


@order_feed_bp.route("/api/order-feed/export", methods=["GET"])
@login_required
def api_order_feed_export():
    """Экспорт ленты заказов в Excel."""
    import urllib.parse
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    try:
        items, _, _, _, _, date_from, date_to, sale_from, sale_to = _load_feed_list(
            refresh=False
        )
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Лента заказов"

    headers = [
        "№ п.п.",
        "Дата заказа",
        "Дата выкупа",
        "Статус",
        "Баркод",
        "Товар",
        "К перечислению",
        "Уникальный ID (srid)",
    ]
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    for idx, it in enumerate(items, 1):
        sale = it.get("sale") or {}
        sold_display = it.get("sold_at_display") or sale.get("sold_at") or ""
        for_pay = sale.get("for_pay")
        product = it.get("name") or ""
        qty = it.get("qty") or 1
        if qty and qty != 1:
            product = f"{qty} шт. · {product}"
        barcode = str(it.get("barcode") or "").strip()

        row = idx + 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=it.get("datetime_display") or "")
        ws.cell(row=row, column=3, value=sold_display if sold_display != "—" else "")
        ws.cell(row=row, column=4, value=it.get("status_label") or "")
        ws.cell(row=row, column=5, value=barcode)
        ws.cell(row=row, column=6, value=product)
        cell_pay = ws.cell(row=row, column=7, value=for_pay if for_pay is not None else "")
        if for_pay is not None:
            cell_pay.number_format = "#,##0.00"
        ws.cell(row=row, column=8, value=it.get("srid") or "")

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    parts = ["Лента_заказов"]
    if date_from and date_to:
        parts.append(f"заказы_{date_from}_{date_to}")
    if sale_from and sale_to:
        parts.append(f"выкупы_{sale_from}_{sale_to}")
    parts.append(datetime.now().strftime("%d.%m.%Y_%H_%M"))
    filename = "_".join(parts) + ".xlsx"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    encoded = urllib.parse.quote(filename.encode("utf-8"))
    return out.getvalue(), 200, {
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Content-Disposition": f"attachment; filename*=UTF-8''{encoded}",
    }


@order_feed_bp.route("/api/order-feed/<path:order_id>", methods=["GET"])
@login_required
def api_order_feed_detail(order_id: str):
    """Детали одного заказа (по srid или gNumber) за указанный период."""
    order_range = _parse_optional_range("date_from", "date_to")
    sale_range = _parse_optional_range("sale_from", "sale_to")
    if not order_range and not sale_range:
        order_range = default_date_range(7)

    orders_from, orders_to, sales_from, sales_to = _resolve_windows(order_range, sale_range)
    user_id = current_user.id
    orders, _ = collect_orders_from_period_cache(user_id, orders_from, orders_to)
    sales = collect_sales_from_period_cache(user_id, sales_from, sales_to)
    items = build_feed_items(user_id, orders, sales, record_history=False)

    needle = str(order_id).strip()
    found = None
    for it in items:
        if it.get("srid") == needle or it.get("gnumber") == needle or it.get("id") == needle:
            found = it
            break

    if not found:
        # расширим поиск: все дни кэша за 90 дней (+ запас на продажи)
        wide_from, wide_to = default_date_range(90)
        if sale_range:
            wide_to = max(wide_to, sale_range[1])
        elif order_range:
            wide_to = max(wide_to, order_range[1])
        orders_w, _ = collect_orders_from_period_cache(user_id, wide_from, wide_to)
        sales_w = collect_sales_from_period_cache(
            user_id, wide_from, extend_iso_date(wide_to, SALES_LOOKUP_EXTRA_DAYS)
        )
        items_w = build_feed_items(user_id, orders_w, sales_w, record_history=False)
        for it in items_w:
            if it.get("srid") == needle or it.get("gnumber") == needle or it.get("id") == needle:
                found = it
                break

    if not found:
        return jsonify({"error": "Заказ не найден", "order_id": needle}), 404

    return jsonify({"item": found})
