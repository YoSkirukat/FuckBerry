# -*- coding: utf-8 -*-
"""Blueprint для поставок FBS (список и состав поставок)."""

import base64
import io
import re
import requests
from datetime import datetime
from flask import Blueprint, jsonify, request, render_template, send_file
from flask_login import login_required, current_user
from utils.wb_token import effective_wb_api_token
from typing import Dict, Any, List

from utils.api import get_with_retry
from utils.cache import (
    load_fbs_supplies_cache,
    save_fbs_supplies_cache,
    load_products_cache,
    load_seller_info_cache_for_user,
)
from utils.api import fetch_all_cards
from utils.constants import (
    FBS_SUPPLIES_LIST_URL,
    FBS_SUPPLY_ORDERS_URL,
    FBS_SUPPLY_ORDERS_IDS_URL_V2,
    FBS_SUPPLY_ORDERS_IDS_URL_V3,
    FBS_SUPPLY_ADD_ORDERS_URL,
    FBS_ORDERS_URL,
    FBS_ORDERS_STICKERS_URL,
    FBS_SUPPLY_BARCODE_URL,
    FBS_SUPPLY_DELIVER_URL,
    FBS_SUPPLY_TRBX_URL,
    FBS_SUPPLY_TRBX_STICKERS_URL,
    MOSCOW_TZ,
)
from utils.helpers import parse_wb_datetime, to_moscow

fbs_supplies_bp = Blueprint("fbs_supplies", __name__)


def _wb_auth_headers(token: str) -> list[dict[str, str]]:
    return [{"Authorization": token}, {"Authorization": f"Bearer {token}"}]


def _fetch_fbs_order_stickers(
    token: str,
    order_ids: list[Any],
    sticker_type: str = "png",
    width: int = 58,
    height: int = 40,
) -> tuple[list[dict[str, Any]], str | None]:
    """Загружает стикеры сборочных заданий (до 100 ID за запрос)."""
    ids: list[int] = []
    for oid in order_ids:
        try:
            ids.append(int(oid))
        except Exception:
            continue
    ids = list(dict.fromkeys(ids))
    if not ids:
        return [], None

    params = {"type": sticker_type, "width": str(width), "height": str(height)}
    all_stickers: list[dict[str, Any]] = []
    last_err: str | None = None

    for i in range(0, len(ids), 100):
        chunk = ids[i : i + 100]
        got_chunk = False
        for hdrs in _wb_auth_headers(token):
            try:
                req_hdrs = dict(hdrs)
                req_hdrs["Content-Type"] = "application/json"
                resp = requests.post(
                    FBS_ORDERS_STICKERS_URL,
                    headers=req_hdrs,
                    params=params,
                    json={"orders": chunk},
                    timeout=60,
                )
                if resp.status_code == 200:
                    data = resp.json() if resp.content else {}
                    all_stickers.extend(data.get("stickers") or [])
                    got_chunk = True
                    break
                last_err = f"HTTP {resp.status_code}: {resp.text[:300]}"
            except Exception as exc:
                last_err = str(exc)
        if not got_chunk:
            break

    return all_stickers, last_err


def _fetch_fbs_supply_barcode(
    token: str,
    supply_id: str,
    sticker_type: str = "png",
) -> tuple[dict[str, Any] | None, str | None]:
    """QR-код поставки (WB-GI-…), для отгрузки на склад WB."""
    url = FBS_SUPPLY_BARCODE_URL.replace("{supplyId}", str(supply_id))
    last_err: str | None = None
    for hdrs in _wb_auth_headers(token):
        try:
            resp = requests.get(url, headers=hdrs, params={"type": sticker_type}, timeout=30)
            if resp.status_code == 200:
                data = resp.json() if resp.content else {}
                if isinstance(data, dict) and data.get("file"):
                    return data, None
                return None, "Пустой ответ API"
            last_err = f"HTTP {resp.status_code}: {resp.text[:300]}"
        except Exception as exc:
            last_err = str(exc)
    return None, last_err


def _generate_qr_png_base64(text: str) -> str:
    """Fallback: генерирует PNG QR-код из текста (если WB API недоступен)."""
    try:
        import qrcode  # type: ignore

        img = qrcode.make(str(text))
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return base64.b64encode(buf.getvalue()).decode("ascii")
    except Exception:
        return ""


def _resolve_supply_barcode(
    token: str,
    supply_id: str,
    sticker_type: str = "png",
) -> dict[str, Any] | None:
    """Возвращает QR поставки: сначала из WB API, иначе локальная генерация по ID."""
    data, _err = _fetch_fbs_supply_barcode(token, supply_id, sticker_type=sticker_type)
    if data and data.get("file"):
        return {
            "code": data.get("barcode") or supply_id,
            "file": data.get("file"),
            "source": "wb",
        }
    generated = _generate_qr_png_base64(supply_id)
    if generated:
        return {"code": supply_id, "file": generated, "source": "generated"}
    return None


def _fetch_fbs_supply_trbx_ids(token: str, supply_id: str) -> list[str]:
    """Список ID грузомест поставки."""
    url = FBS_SUPPLY_TRBX_URL.replace("{supplyId}", str(supply_id))
    for hdrs in _wb_auth_headers(token):
        try:
            resp = requests.get(url, headers=hdrs, timeout=30)
            if resp.status_code != 200:
                continue
            data = resp.json() if resp.content else {}
            trbxes = data.get("trbxes") if isinstance(data, dict) else []
            ids: list[str] = []
            for it in trbxes or []:
                if isinstance(it, dict):
                    tid = it.get("id") or it.get("trbxId")
                    if tid:
                        ids.append(str(tid))
                elif it:
                    ids.append(str(it))
            return ids
        except Exception:
            continue
    return []


def _fetch_fbs_trbx_stickers(
    token: str,
    supply_id: str,
    trbx_ids: list[str],
    sticker_type: str = "png",
) -> list[dict[str, Any]]:
    """QR-коды грузомест (для отгрузки в ПВЗ)."""
    if not trbx_ids:
        return []
    url = FBS_SUPPLY_TRBX_STICKERS_URL.replace("{supplyId}", str(supply_id))
    for hdrs in _wb_auth_headers(token):
        try:
            req_hdrs = dict(hdrs)
            req_hdrs["Content-Type"] = "application/json"
            resp = requests.post(
                url,
                headers=req_hdrs,
                params={"type": sticker_type},
                json={"trbxIds": trbx_ids},
                timeout=60,
            )
            if resp.status_code != 200:
                continue
            data = resp.json() if resp.content else {}
            stickers = data.get("stickers") or []
            out: list[dict[str, Any]] = []
            for i, tid in enumerate(trbx_ids):
                st = stickers[i] if i < len(stickers) and isinstance(stickers[i], dict) else {}
                out.append(
                    {
                        "id": tid,
                        "file": st.get("file") or "",
                        "barcode": st.get("barcode") or "",
                    }
                )
            return out
        except Exception:
            continue
    return [{"id": tid, "file": "", "barcode": ""} for tid in trbx_ids]


def _add_fbs_supply_trbx(
    token: str,
    supply_id: str,
    amount: int = 1,
) -> tuple[list[str], str | None]:
    """Добавляет грузоместа в поставку."""
    url = FBS_SUPPLY_TRBX_URL.replace("{supplyId}", str(supply_id))
    amount = max(1, min(int(amount), 1000))
    last_err: str | None = None
    for hdrs in _wb_auth_headers(token):
        try:
            req_hdrs = dict(hdrs)
            req_hdrs["Content-Type"] = "application/json"
            resp = requests.post(url, headers=req_hdrs, json={"amount": amount}, timeout=30)
            if resp.status_code in (200, 201):
                data = resp.json() if resp.content else {}
                ids = data.get("trbxIds") or []
                return [str(x) for x in ids], None
            last_err = f"HTTP {resp.status_code}: {resp.text[:300]}"
        except Exception as exc:
            last_err = str(exc)
    return [], last_err


def _load_trbx_items(token: str, supply_id: str) -> list[dict[str, Any]]:
    ids = _fetch_fbs_supply_trbx_ids(token, supply_id)
    if not ids:
        return []
    return _fetch_fbs_trbx_stickers(token, supply_id, ids)


def _char_value_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(x) for x in value if x is not None and str(x).strip())
    return str(value).strip()


def _extract_card_color(card: dict[str, Any]) -> str:
    for ch in card.get("characteristics") or []:
        if not isinstance(ch, dict):
            continue
        name = str(ch.get("name") or "").strip().lower()
        if "цвет" in name and "количество" not in name:
            return _char_value_to_str(ch.get("value"))
    return ""


def _resolve_variant_from_card(
    card: dict[str, Any],
    chrt_id: Any = None,
    barcode: str = "",
) -> dict[str, str]:
    brand = str(card.get("brand") or "").strip()
    title = str(card.get("title") or card.get("name") or "").strip()
    article = str(card.get("vendorCode") or card.get("supplierArticle") or "").strip()
    color = _extract_card_color(card)
    size = ""
    bc = str(barcode or "").strip()
    chrt: int | None = None
    try:
        chrt = int(chrt_id) if chrt_id is not None else None
    except Exception:
        chrt = None

    for s in card.get("sizes") or []:
        if not isinstance(s, dict):
            continue
        sid = s.get("chrtID") or s.get("chrtId")
        skus = [str(x) for x in (s.get("skus") or s.get("barcodes") or []) if x]
        matched = False
        if chrt is not None and sid is not None:
            try:
                matched = int(sid) == chrt
            except Exception:
                matched = False
        if not matched and bc and bc in skus:
            matched = True
        if matched:
            size = str(s.get("techSize") or s.get("wbSize") or "").strip()
            if not bc and skus:
                bc = skus[0]
            break

    if not size:
        sizes = card.get("sizes") or []
        if sizes and isinstance(sizes[0], dict):
            size = str(sizes[0].get("techSize") or sizes[0].get("wbSize") or "").strip()

    return {
        "brand": brand,
        "title": title,
        "article": article,
        "color": color,
        "size": size,
        "barcode": bc,
    }


def _build_card_lookup(token: str, nm_ids: set[int]) -> dict[int, dict[str, Any]]:
    if not token or not nm_ids:
        return {}
    lookup: dict[int, dict[str, Any]] = {}
    try:
        for card in fetch_all_cards(token, page_limit=100):
            if not isinstance(card, dict):
                continue
            nmid = card.get("nmID") or card.get("nmId") or card.get("nm")
            if nmid is None:
                continue
            try:
                nmid_i = int(nmid)
            except Exception:
                continue
            if nmid_i in nm_ids:
                lookup[nmid_i] = card
    except Exception:
        pass
    return lookup


def _generate_barcode_png_base64(code: str) -> str:
    """Генерирует PNG штрихкода (EAN-13 или Code128) в base64."""
    pil = _generate_barcode_pil(code)
    if pil is None:
        return ""
    buf = io.BytesIO()
    pil.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode("ascii")


# Размер стикера WB: 58×40 мм = 580×400 px при 254 DPI.
_STICKER_W = 580
_STICKER_H = 400
_PRODUCT_LABEL_SCALE = 1
_PRODUCT_LABEL_W = _STICKER_W * _PRODUCT_LABEL_SCALE
_PRODUCT_LABEL_H = _STICKER_H * _PRODUCT_LABEL_SCALE
_PRODUCT_LABEL_DPI = 254
_BARCODE_WRITER_DPI = 300


def _wrap_label_lines_plain(text: str, max_chars: int = 38, max_lines: int = 3) -> list[str]:
    """Перенос строк для HTML/PDF без PIL."""
    words = str(text or "").split()
    if not words:
        return []
    lines: list[str] = []
    current = words[0]
    for word in words[1:]:
        trial = f"{current} {word}"
        if len(trial) <= max_chars:
            current = trial
        else:
            lines.append(current)
            current = word
            if len(lines) >= max_lines:
                break
    if len(lines) < max_lines:
        lines.append(current)
    if len(lines) > max_lines:
        lines = lines[:max_lines]
    if len(lines) == max_lines:
        joined = " ".join(lines)
        rest = " ".join(words[len(joined.split()):]) if words else ""
        if rest:
            last = lines[-1]
            while len(last) > 3 and len(last + "…") > max_chars:
                last = last[:-1]
            lines[-1] = last + "…"
    return lines


_BARCODE_MODULE_WIDTH = 0.35  # +25% к прежним 0.28 для более широкого штрихкода


def _generate_barcode_svg_inline(code: str) -> str:
    """Векторный штрихкод (SVG) для чёткой печати текста и полос."""
    digits = re.sub(r"\D", "", str(code or ""))
    if len(digits) < 8:
        return ""
    try:
        from barcode import EAN13  # type: ignore
        from barcode.codex import Code128  # type: ignore
        from barcode.writer import SVGWriter  # type: ignore

        writer = SVGWriter()
        writer.set_options(
            {
                "module_height": 6,
                "module_width": _BARCODE_MODULE_WIDTH,
                "quiet_zone": 1.5,
                "font_size": 0,
                "text_distance": 0,
            }
        )
        buf = io.BytesIO()
        opts = {"write_text": False}
        if len(digits) >= 12:
            EAN13(digits[:12], writer=writer).write(buf, options=opts)
        else:
            Code128(digits, writer=writer).write(buf, options=opts)
        svg = buf.getvalue().decode("utf-8")
        svg = re.sub(r'width="[^"]*"', 'width="100%"', svg, count=1)
        svg = re.sub(r'height="[^"]*"', 'height="auto"', svg, count=1)
        return svg.replace("<svg ", '<svg preserveAspectRatio="xMidYMid meet" ', 1)
    except Exception:
        return ""


def _build_product_label_view(label: dict[str, Any]) -> dict[str, Any]:
    """Данные этикетки товара для HTML/PDF (векторный штрихкод + текст, как в ЛК WB)."""
    barcode = str(label.get("barcode") or "").strip()
    title = str(label.get("title") or "").strip()
    article = _truncate_text(label.get("article") or "", 44)
    return {
        "barcode": barcode,
        "seller": str(label.get("seller") or "").strip(),
        "title": title,
        "title_lines": _wrap_label_lines_plain(title, max_chars=34, max_lines=2),
        "brand": str(label.get("brand") or "").strip(),
        "color": _truncate_text(label.get("color") or "", 36),
        "size": str(label.get("size") or "").strip(),
        "article": article,
        "article_lines": _wrap_label_lines_plain(article, max_chars=30, max_lines=2),
        "barcode_svg": _generate_barcode_svg_inline(barcode),
    }


_REPORTLAB_FONTS_READY = False
_REPORTLAB_FONT = "Helvetica"
_REPORTLAB_FONT_BOLD = "Helvetica-Bold"


def _ensure_reportlab_fonts() -> tuple[str, str]:
    global _REPORTLAB_FONTS_READY, _REPORTLAB_FONT, _REPORTLAB_FONT_BOLD
    if _REPORTLAB_FONTS_READY:
        return _REPORTLAB_FONT, _REPORTLAB_FONT_BOLD
    from reportlab.pdfbase import pdfmetrics  # type: ignore
    from reportlab.pdfbase.ttfonts import TTFont  # type: ignore

    candidates = [
        (
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        ),
        (
            "/usr/share/fonts/TTF/DejaVuSans.ttf",
            "/usr/share/fonts/TTF/DejaVuSans-Bold.ttf",
        ),
    ]
    for regular_path, bold_path in candidates:
        try:
            pdfmetrics.registerFont(TTFont("DejaVu", regular_path))
            pdfmetrics.registerFont(TTFont("DejaVu-Bold", bold_path))
            _REPORTLAB_FONT = "DejaVu"
            _REPORTLAB_FONT_BOLD = "DejaVu-Bold"
            break
        except Exception:
            continue
    _REPORTLAB_FONTS_READY = True
    return _REPORTLAB_FONT, _REPORTLAB_FONT_BOLD


def _pdf_font_height(font_name: str, size: float) -> tuple[float, float]:
    from reportlab.pdfbase.pdfmetrics import getAscent, getDescent  # type: ignore

    return getAscent(font_name) / 1000.0 * size, getDescent(font_name) / 1000.0 * size


def _pdf_baseline_below(
    prev_baseline: float,
    prev_font: str,
    prev_size: float,
    gap: float,
    next_font: str,
    next_size: float,
) -> float:
    """Baseline следующей строки под предыдущей (PDF: y растёт вверх)."""
    _, descent = _pdf_font_height(prev_font, prev_size)
    ascent, _ = _pdf_font_height(next_font, next_size)
    return prev_baseline - descent - gap - ascent


def _pdf_draw_centered(
    c: Any,
    text: str,
    baseline: float,
    page_w: float,
    font: str,
    size: float,
) -> float:
    if not text:
        return baseline
    c.setFont(font, size)
    c.drawCentredString(page_w / 2, baseline, text)
    return baseline


def _draw_pdf_attr_line(
    c: Any,
    label: str,
    value: str,
    baseline: float,
    page_w: float,
    font: str,
    font_size: float = 6.5,
    gap_after: float = 1.1,
) -> float:
    from reportlab.lib.units import mm  # type: ignore
    from reportlab.pdfbase.pdfmetrics import stringWidth  # type: ignore

    text = f"{label}: "
    val = str(value or "")
    tw_label = stringWidth(text, font, font_size)
    tw_val = stringWidth(val, font, font_size) if val else 0
    x = (page_w - tw_label - tw_val) / 2
    c.setFont(font, font_size)
    c.drawString(x, baseline, text)
    if val:
        c.drawString(x + tw_label, baseline, val)
    return _pdf_baseline_below(baseline, font, font_size, gap_after * mm, font, font_size)


def _draw_product_label_pdf_page(c: Any, label: dict[str, Any], page_w: float, page_h: float) -> None:
    """Рисует этикетку товара вектором на странице PDF 58×40 мм."""
    from reportlab.graphics import renderPDF  # type: ignore
    from reportlab.graphics.barcode import createBarcodeDrawing  # type: ignore
    from reportlab.lib.units import mm  # type: ignore

    font, _font_bold = _ensure_reportlab_fonts()
    barcode = str(label.get("barcode") or "").strip()
    seller = str(label.get("seller") or "").strip()
    title_lines = label.get("title_lines") or _wrap_label_lines_plain(label.get("title"), max_lines=2)
    brand = str(label.get("brand") or "").strip()
    color = str(label.get("color") or "").strip()
    size_val = str(label.get("size") or "").strip()
    article = str(label.get("article") or "").strip()

    num_size = 8.0
    seller_size = 7.0
    title_size = 7.5
    attr_size = 6.5

    y_top = page_h - 1.5 * mm
    bc_bottom = y_top - 11 * mm

    if barcode and len(re.sub(r"\D", "", barcode)) >= 12:
        try:
            bc = createBarcodeDrawing(
                "EAN13",
                value=barcode,
                barWidth=_BARCODE_MODULE_WIDTH * mm,
                barHeight=11 * mm,
                humanReadable=False,
            )
            scale = min(1.0, (page_w - 4 * mm) / bc.width)
            if scale < 1.0:
                bc.width *= scale
                bc.height *= scale
            bx = (page_w - bc.width) / 2
            bc_bottom = y_top - bc.height
            renderPDF.draw(bc, c, bx, bc_bottom)
        except Exception:
            bc_bottom = y_top - 11 * mm

    num_ascent, _ = _pdf_font_height(font, num_size)
    baseline = bc_bottom - 1.2 * mm - num_ascent
    baseline = _pdf_draw_centered(c, barcode, baseline, page_w, font, num_size)
    prev_font, prev_size = font, num_size

    if seller:
        baseline = _pdf_baseline_below(baseline, prev_font, prev_size, 4 * mm, font, seller_size)
        baseline = _pdf_draw_centered(c, seller, baseline, page_w, font, seller_size)
        prev_font, prev_size = font, seller_size

    for i, line in enumerate(title_lines):
        if not line:
            continue
        gap = 1.5 * mm if i == 0 else 1.0 * mm
        baseline = _pdf_baseline_below(baseline, prev_font, prev_size, gap, font, title_size)
        baseline = _pdf_draw_centered(c, line, baseline, page_w, font, title_size)
        prev_font, prev_size = font, title_size

    baseline = _pdf_baseline_below(baseline, prev_font, prev_size, 1.0 * mm, font, attr_size)
    baseline = _draw_pdf_attr_line(c, "Бренд", brand, baseline, page_w, font, attr_size)
    baseline = _draw_pdf_attr_line(c, "Цвет", color, baseline, page_w, font, attr_size)
    baseline = _draw_pdf_attr_line(c, "Размер", size_val, baseline, page_w, font, attr_size)

    article_lines = _wrap_label_lines_plain(article, max_chars=30, max_lines=2)
    if len(article_lines) <= 1:
        _draw_pdf_attr_line(c, "Артикул", article_lines[0] if article_lines else "", baseline, page_w, font, attr_size, gap_after=0)
    else:
        baseline = _draw_pdf_attr_line(c, "Артикул", article_lines[0], baseline, page_w, font, attr_size, gap_after=0.7)
        baseline = _pdf_baseline_below(baseline, font, attr_size, 0.6 * mm, font, attr_size)
        _pdf_draw_centered(c, article_lines[1], baseline, page_w, font, attr_size)


def _build_tape_pdf(
    pairs: list[dict[str, Any]],
    *,
    include_barcodes: bool = True,
    include_stickers: bool = True,
) -> io.BytesIO:
    """PDF ленты: векторные этикетки товаров + растровые стикеры WB."""
    from reportlab.lib.units import mm  # type: ignore
    from reportlab.lib.utils import ImageReader  # type: ignore
    from reportlab.pdfgen import canvas  # type: ignore

    page_w, page_h = 58 * mm, 40 * mm
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(page_w, page_h))
    pages = 0

    for pair in pairs:
        if not isinstance(pair, dict):
            continue
        label = pair.get("label")
        if include_barcodes and isinstance(label, dict) and label.get("barcode"):
            _draw_product_label_pdf_page(c, label, page_w, page_h)
            c.showPage()
            pages += 1
        sticker_file = pair.get("sticker_file")
        if include_stickers and sticker_file:
            try:
                img_bytes = _optimize_image_for_pdf(base64.b64decode(str(sticker_file)))
                c.drawImage(
                    ImageReader(io.BytesIO(img_bytes)),
                    0,
                    0,
                    width=page_w,
                    height=page_h,
                    preserveAspectRatio=True,
                    anchor="sw",
                )
                c.showPage()
                pages += 1
            except Exception:
                pass

    if pages == 0:
        raise ValueError("Нет стикеров для PDF")

    c.save()
    buf.seek(0)
    return buf


def _merge_pdf_buffers(buffers: list[io.BytesIO]) -> io.BytesIO:
    from pypdf import PdfReader, PdfWriter  # type: ignore

    writer = PdfWriter()
    for buf in buffers:
        buf.seek(0)
        reader = PdfReader(buf)
        for page in reader.pages:
            writer.add_page(page)
    out = io.BytesIO()
    writer.write(out)
    out.seek(0)
    return out


def _truthy_arg(value: str | None) -> bool:
    return str(value or "").strip().lower() in ("1", "true", "yes", "on")


def _build_fbs_print_pdf(
    ctx: dict[str, Any],
    *,
    qr: bool,
    trbx: bool,
    stickers: bool,
    barcodes: bool,
) -> io.BytesIO:
    """Собирает PDF для печати по выбранным блокам (тот же рендер, что и «Сохранить в PDF»)."""
    parts: list[io.BytesIO] = []

    if qr:
        barcode = ctx.get("barcode") or {}
        if isinstance(barcode, dict) and barcode.get("file"):
            parts.append(_build_images_pdf([barcode["file"]]))

    if trbx:
        trbx_images = [
            item["file"]
            for item in (ctx.get("trbx_items") or [])
            if isinstance(item, dict) and item.get("file")
        ]
        if trbx_images:
            parts.append(_build_images_pdf(trbx_images))

    if stickers or barcodes:
        pairs = ctx.get("order_tape_pairs") or []
        if pairs:
            parts.append(
                _build_tape_pdf(
                    pairs,
                    include_barcodes=barcodes,
                    include_stickers=stickers,
                )
            )

    if not parts:
        raise ValueError("Нет данных для печати")

    if len(parts) == 1:
        return parts[0]
    return _merge_pdf_buffers(parts)


def _to_print_bitmap(img: Any) -> Any:
    """Чистый ч/б без полутонов — лучше читается на термопринтере."""
    from PIL import Image  # type: ignore

    gray = img.convert("L")
    return gray.point(lambda p: 255 if p > 160 else 0, mode="1")


def _encode_sticker_png(img: Any, dpi: int = _PRODUCT_LABEL_DPI) -> bytes:
    """Сжатый 1-bit PNG для стикера 58×40 мм."""
    from PIL import Image  # type: ignore

    if img.mode != "1":
        img = _to_print_bitmap(img)
    if img.size != (_STICKER_W, _STICKER_H):
        img = img.resize((_STICKER_W, _STICKER_H), Image.Resampling.NEAREST)
    buf = io.BytesIO()
    img.save(buf, format="PNG", dpi=(dpi, dpi), optimize=True, compress_level=9)
    return buf.getvalue()


def _optimize_image_for_pdf(image_bytes: bytes) -> bytes:
    """Нормализует и сжимает изображение перед вставкой в PDF."""
    from PIL import Image  # type: ignore

    try:
        img = Image.open(io.BytesIO(image_bytes))
    except Exception:
        return image_bytes

    sticker_ratio = _STICKER_W / _STICKER_H
    img_ratio = img.width / max(img.height, 1)
    is_sticker_like = abs(img_ratio - sticker_ratio) < 0.2

    if is_sticker_like and (img.width != _STICKER_W or img.height != _STICKER_H):
        img = img.resize((_STICKER_W, _STICKER_H), Image.Resampling.NEAREST)

    return _encode_sticker_png(img)


def _generate_barcode_pil(
    code: str,
    target_width: int = 500,
    min_height: int | None = None,
) -> Any | None:
    """Генерирует PIL-изображение штрихкода в высоком разрешении."""
    digits = re.sub(r"\D", "", str(code or ""))
    if len(digits) < 8:
        return None
    try:
        from barcode import EAN13  # type: ignore
        from barcode.codex import Code128  # type: ignore
        from barcode.writer import ImageWriter  # type: ignore
        from PIL import Image  # type: ignore

        writer = ImageWriter()
        writer.set_options(
            {
                "module_width": 0.4,
                "module_height": 16,
                "font_size": 0,
                "text_distance": 0,
                "quiet_zone": 1.5,
                "dpi": _BARCODE_WRITER_DPI,
            }
        )
        buf = io.BytesIO()
        opts = {"write_text": False}
        if len(digits) >= 12:
            EAN13(digits[:12], writer=writer).write(buf, options=opts)
        else:
            Code128(digits, writer=writer).write(buf, options=opts)
        img = Image.open(buf).convert("RGB")
        img = _to_print_bitmap(img)

        if min_height and img.height < min_height:
            new_h = min_height
            new_w = max(1, round(img.width * new_h / img.height))
            if new_w > target_width:
                new_w = target_width
                new_h = max(1, round(img.height * target_width / img.width))
            img = img.resize((new_w, new_h), Image.Resampling.NEAREST)
        elif img.width != target_width:
            new_h = max(1, round(img.height * target_width / img.width))
            img = img.resize((target_width, new_h), Image.Resampling.NEAREST)
        return img.convert("RGB") if img.mode == "1" else img
    except Exception:
        return None


def _load_product_label_fonts(scale: int) -> dict[str, Any]:
    """Шрифты пропорционально эталону WB (580×400)."""
    from PIL import ImageFont  # type: ignore

    regular_size = max(24, round(24 * scale))
    title_size = max(27, round(27 * scale))
    num_size = max(38, round(38 * scale))
    bold_size = max(26, round(26 * scale))
    candidates = [
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("/usr/share/fonts/TTF/DejaVuSans.ttf", "/usr/share/fonts/TTF/DejaVuSans-Bold.ttf"),
    ]
    for regular, bold in candidates:
        try:
            return {
                "regular": ImageFont.truetype(regular, regular_size),
                "title": ImageFont.truetype(regular, title_size),
                "bold": ImageFont.truetype(bold, bold_size),
                "num": ImageFont.truetype(regular, num_size),
                "stroke": max(1, scale // 3),
            }
        except Exception:
            continue
    default = ImageFont.load_default()
    return {"regular": default, "title": default, "bold": default, "num": default, "stroke": 0}


def _draw_text_centered(
    draw: Any,
    x_center: int,
    y: int,
    text: str,
    font: Any,
    stroke: int = 0,
) -> int:
    if not text:
        return y
    bbox = draw.textbbox((0, 0), text, font=font, stroke_width=stroke)
    tw = int(bbox[2] - bbox[0])
    draw.text(
        (x_center - tw // 2, y),
        text,
        font=font,
        fill="black",
        stroke_width=stroke,
        stroke_fill="black",
    )
    return y + int(bbox[3] - bbox[1])


def _text_width(draw: Any, text: str, font: Any, stroke: int = 0) -> int:
    if not text:
        return 0
    try:
        bbox = draw.textbbox((0, 0), text, font=font, stroke_width=stroke)
        return int(bbox[2] - bbox[0])
    except Exception:
        return len(text) * 8


def _wrap_text_lines(
    draw: Any,
    text: str,
    font: Any,
    max_width: int,
    max_lines: int = 2,
    stroke: int = 0,
) -> list[str]:
    words = str(text or "").split()
    if not words:
        return []
    lines: list[str] = []
    current = words[0]
    for word in words[1:]:
        trial = f"{current} {word}"
        if _text_width(draw, trial, font, stroke) <= max_width:
            current = trial
        else:
            lines.append(current)
            current = word
            if len(lines) >= max_lines:
                break
    if len(lines) < max_lines:
        lines.append(current)
    if len(lines) > max_lines:
        lines = lines[:max_lines]
    if len(lines) == max_lines and len(words) > len(" ".join(lines).split()):
        last = lines[-1]
        while last and _text_width(draw, last + "…", font, stroke) > max_width:
            last = last[:-1]
        lines[-1] = (last + "…") if last else "…"
    return lines


def _draw_centered_lines(
    draw: Any,
    lines: list[str],
    y: int,
    font: Any,
    canvas_w: int,
    line_gap: int = 2,
    stroke: int = 0,
) -> int:
    for line in lines:
        if not line:
            continue
        y = _draw_text_centered(draw, canvas_w // 2, y, line, font, stroke=stroke)
        y += line_gap
    return y


def _draw_attr_line(
    draw: Any,
    label: str,
    value: str,
    y: int,
    fonts: dict,
    canvas_w: int,
) -> int:
    font = fonts["regular"]
    bold = fonts["bold"]
    stroke = fonts.get("stroke", 0)
    text = f"{label}: "
    val = str(value or "")
    tw_label = _text_width(draw, text, font, stroke)
    tw_val = _text_width(draw, val, bold, stroke) if val else 0
    total = tw_label + tw_val
    x = (canvas_w - total) // 2
    draw.text((x, y), text, font=font, fill="black", stroke_width=stroke, stroke_fill="black")
    if val:
        draw.text(
            (x + tw_label, y),
            val,
            font=bold,
            fill="black",
            stroke_width=stroke,
            stroke_fill="black",
        )
    bbox = draw.textbbox((0, 0), text or "Ay", font=font, stroke_width=stroke)
    return y + int(bbox[3] - bbox[1]) + 1


def _truncate_text(text: str, max_len: int = 32) -> str:
    s = str(text or "").strip()
    if len(s) <= max_len:
        return s
    return s[: max_len - 1] + "…"


def _render_product_label_png(label: dict[str, Any]) -> str:
    """Рендерит стикер штрихкода товара 580×400 мм-эквивалент в повышенном разрешении."""
    try:
        from PIL import Image, ImageDraw  # type: ignore
    except Exception:
        return ""

    W, H = _PRODUCT_LABEL_W, _PRODUCT_LABEL_H
    scale = _PRODUCT_LABEL_SCALE
    img = Image.new("RGB", (W, H), "white")
    draw = ImageDraw.Draw(img)
    fonts = _load_product_label_fonts(scale)
    # Отступы от высоты этикетки — иначе на 3× рендере 5*scale даёт всего ~15 px.
    block_gap = int(H * 0.065)
    line_gap = int(H * 0.045)
    pad_x = 10 * scale
    max_text_w = W - pad_x * 2
    y = int(H * 0.02)

    barcode = str(label.get("barcode") or "").strip()
    bc_width = int(W * 0.96)
    bc_img = _generate_barcode_pil(barcode, target_width=bc_width) if barcode else None
    if bc_img is not None:
        new_h = max(1, bc_img.height // 2)
        bc_img = bc_img.resize((bc_width, new_h), Image.Resampling.NEAREST)
        x = (W - bc_img.width) // 2
        img.paste(bc_img, (x, y))
        y += bc_img.height + block_gap

    if barcode:
        y = _draw_text_centered(draw, W // 2, y, barcode, fonts["num"], stroke=0)
        y += block_gap

    seller = str(label.get("seller") or "").strip()
    if seller:
        y = _draw_centered_lines(draw, [seller], y, fonts["regular"], W, line_gap=line_gap, stroke=0)
        y += block_gap

    article = str(label.get("article") or "").strip()
    if article:
        article_lines = _wrap_text_lines(
            draw, article, fonts["title"], max_text_w, max_lines=3, stroke=0
        )
        _draw_centered_lines(draw, article_lines, y, fonts["title"], W, line_gap=line_gap, stroke=0)

    img = _to_print_bitmap(img)
    png_bytes = _encode_sticker_png(img)

    return base64.b64encode(png_bytes).decode("ascii")


def _load_seller_organization_name(user_id: int) -> str:
    cached = load_seller_info_cache_for_user(user_id)
    if not cached:
        return ""
    name = cached.get("organization_name")
    if not name and isinstance(cached.get("seller_info"), dict):
        name = cached["seller_info"].get("name")
    return str(name or "").strip()


def _build_product_barcode_labels(
    raw_items: list[dict[str, Any]],
    norm_items: list[dict[str, Any]],
    token: str,
    user_id: int,
) -> list[dict[str, Any]]:
    """Собирает данные для печати штрихкодов товаров (как в ЛК WB)."""
    seller = _load_seller_organization_name(user_id)
    norm_by_id: dict[Any, dict[str, Any]] = {}
    for it in norm_items:
        oid = it.get("id")
        if oid is not None:
            norm_by_id[oid] = it

    prod_cached = load_products_cache() or {}
    by_nm_prod: dict[int, dict[str, Any]] = {}
    for it in prod_cached.get("items") or []:
        nmv = it.get("nm_id") or it.get("nmID")
        if nmv:
            try:
                by_nm_prod[int(nmv)] = it
            except Exception:
                continue

    nm_ids: set[int] = set()
    for it in raw_items:
        nm = it.get("nmId") or it.get("nmID")
        if nm is None:
            continue
        try:
            nm_ids.add(int(nm))
        except Exception:
            continue

    card_by_nm = _build_card_lookup(token, nm_ids)
    labels: list[dict[str, Any]] = []

    for raw in raw_items:
        if not isinstance(raw, dict):
            continue
        oid = raw.get("id") or raw.get("orderId")
        norm = norm_by_id.get(oid, {})
        barcode = str(norm.get("barcode") or "").strip()
        skus = raw.get("skus") or []
        if isinstance(skus, list) and skus and not barcode:
            barcode = str(skus[0]).strip()

        nm = raw.get("nmId") or raw.get("nmID")
        chrt = raw.get("chrtId") or raw.get("chrtID")
        card = None
        if nm is not None:
            try:
                card = card_by_nm.get(int(nm))
            except Exception:
                card = None

        brand = color = size = title = article = ""
        if card:
            info = _resolve_variant_from_card(card, chrt_id=chrt, barcode=barcode)
            brand = info["brand"]
            color = info["color"]
            size = info["size"]
            title = info["title"]
            article = info["article"] or str(norm.get("article") or raw.get("article") or "").strip()
            if info["barcode"]:
                barcode = info["barcode"]
        else:
            hit = None
            if nm is not None:
                try:
                    hit = by_nm_prod.get(int(nm))
                except Exception:
                    hit = None
            if hit:
                title = str(hit.get("name") or "").strip()
                article = str(hit.get("supplier_article") or norm.get("article") or "").strip()
            else:
                fallback = str(norm.get("article") or raw.get("article") or "").strip()
                title = fallback
                article = fallback

        if not article:
            article = str(norm.get("article") or raw.get("article") or "").strip()

        label_data = {
            "order_id": oid,
            "barcode": barcode,
            "seller": seller,
            "title": title,
            "brand": brand,
            "color": color,
            "size": size,
            "article": article,
        }
        label_data.update(_build_product_label_view(label_data))
        labels.append(label_data)

    return labels


def _build_order_tape_pairs(
    product_labels: list[dict[str, Any]],
    stickers: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Пары «штрихкод товара + стикер сборки» по order_id (как лента в ЛК WB)."""
    sticker_by_id: dict[str, dict[str, Any]] = {}
    for sticker in stickers:
        oid = sticker.get("order_id")
        if oid is None:
            continue
        sticker_by_id[str(oid)] = sticker

    pairs: list[dict[str, Any]] = []
    used_sticker_ids: set[str] = set()

    for label in product_labels:
        oid = label.get("order_id")
        oid_key = str(oid) if oid is not None else ""
        sticker = sticker_by_id.get(oid_key)
        if sticker:
            used_sticker_ids.add(oid_key)
        pairs.append(
            {
                "order_id": oid,
                "label": {
                    "barcode": label.get("barcode") or "",
                    "seller": label.get("seller") or "",
                    "title": label.get("title") or "",
                    "title_lines": label.get("title_lines") or [],
                    "brand": label.get("brand") or "",
                    "color": label.get("color") or "",
                    "size": label.get("size") or "",
                    "article": label.get("article") or "",
                    "article_lines": label.get("article_lines") or [],
                    "barcode_svg": label.get("barcode_svg") or "",
                }
                if label.get("barcode")
                else None,
                "sticker_file": (sticker or {}).get("file") or "",
            }
        )

    for sticker in stickers:
        oid = sticker.get("order_id")
        oid_key = str(oid) if oid is not None else ""
        if oid_key in used_sticker_ids:
            continue
        pairs.append(
            {
                "order_id": oid,
                "label": None,
                "sticker_file": sticker.get("file") or "",
            }
        )

    # Как в листе подбора: одинаковые товары подряд (баркод, затем артикул)
    def _tape_sort_key(pair: dict[str, Any]) -> tuple[str, str]:
        label = pair.get("label") if isinstance(pair.get("label"), dict) else {}
        return (
            str((label or {}).get("barcode") or "").strip(),
            str((label or {}).get("article") or "").strip(),
        )

    pairs.sort(key=_tape_sort_key)
    return pairs


def _safe_supply_filename_part(supply_id: str) -> str:
    return "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in str(supply_id))


def _download_date_label() -> str:
    return datetime.now(MOSCOW_TZ).strftime("%d.%m.%Y")


def _dated_download_name(title: str, ext: str) -> str:
    """Имя файла вида «Заголовок DD.MM.YYYY.ext»."""
    return f"{title} {_download_date_label()}.{ext.lstrip('.')}"


def _build_images_pdf(images_b64: list[str]) -> io.BytesIO:
    """Собирает PDF: одно изображение на страницу (размер страницы = размеру картинки)."""
    import img2pdf  # type: ignore

    payloads: list[bytes] = []
    for item in images_b64:
        raw = str(item or "").strip()
        if not raw:
            continue
        try:
            payloads.append(_optimize_image_for_pdf(base64.b64decode(raw)))
        except Exception:
            continue
    if not payloads:
        raise ValueError("Нет изображений для PDF")

    out = io.BytesIO()
    out.write(img2pdf.convert(payloads))
    out.seek(0)
    return out


def _format_sticker_code(part_a: Any, part_b: Any) -> str:
    a = str(part_a or "").strip()
    b = str(part_b or "").strip()
    if a and b:
        return f"{a} {b}"
    return a or b


def _build_pick_list_rows(ctx: dict[str, Any]) -> list[dict[str, Any]]:
    """Строки листа подбора: фото, артикул, стикер, баркод (по одному на заказ)."""
    raw_by_id: dict[str, dict[str, Any]] = {}
    for raw in ctx.get("raw_items") or []:
        if not isinstance(raw, dict):
            continue
        oid = raw.get("id") or raw.get("orderId")
        if oid is not None:
            raw_by_id[str(oid)] = raw

    sticker_by_id: dict[str, dict[str, Any]] = {}
    for sticker in ctx.get("stickers") or []:
        if not isinstance(sticker, dict):
            continue
        oid = sticker.get("order_id")
        if oid is not None:
            sticker_by_id[str(oid)] = sticker

    rows: list[dict[str, Any]] = []
    for item in ctx.get("items") or []:
        if not isinstance(item, dict):
            continue
        oid = item.get("id")
        oid_key = str(oid) if oid is not None else ""
        raw = raw_by_id.get(oid_key) or {}
        sticker = sticker_by_id.get(oid_key) or {}

        barcode = str(item.get("barcode") or "").strip()
        if not barcode:
            skus = raw.get("skus") or []
            if isinstance(skus, list) and skus:
                barcode = str(skus[0]).strip()

        rows.append(
            {
                "photo": item.get("photo"),
                "article": str(item.get("article") or "").strip(),
                "sticker": _format_sticker_code(sticker.get("part_a"), sticker.get("part_b")),
                "barcode": barcode,
            }
        )

    # Одинаковые товары — подряд (как в ЛК WB)
    rows.sort(key=lambda r: (r.get("barcode") or "", r.get("article") or ""))
    return rows


def _fetch_product_photo_bytes(url: str) -> bytes | None:
    photo_url = str(url or "").strip()
    if not photo_url:
        return None
    try:
        resp = requests.get(photo_url, timeout=12)
        if resp.status_code == 200 and resp.content:
            return resp.content
    except Exception:
        pass
    return None


def _build_pick_list_xlsx(supply_id: str, rows: list[dict[str, Any]]) -> io.BytesIO:
    """Собирает лист подбора в формате, близком к ЛК WB."""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Лист подбора"

    today_str = datetime.now(MOSCOW_TZ).strftime("%d.%m.%Y")
    purple = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws["A1"] = f"Дата: {today_str}"
    ws["A1"].font = Font(size=11)

    ws.merge_cells("A2:E2")
    title_cell = ws["A2"]
    title_cell.value = f"Лист подбора {supply_id}"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = purple
    title_cell.alignment = center
    ws.row_dimensions[2].height = 28

    ws["A4"] = f"Количество товаров: {len(rows)}"
    ws["A4"].font = Font(size=11)

    headers = ["Фото", "Артикул продавца", "Стикер", "Баркод", "Коробка"]
    header_row = 5
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = Font(bold=True, size=11)
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    ws.column_dimensions["A"].width = 13.5
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14

    # Размер фото в Excel: 2,5 × 2 см (пропорции как в ЛК WB)
    photo_width_px = round(2.5 / 2.54 * 96)
    photo_height_px = round(2.0 / 2.54 * 96)
    row_height_pt = 2.0 * 28.3465
    data_start = header_row + 1
    for idx, row in enumerate(rows):
        r = data_start + idx
        ws.row_dimensions[r].height = row_height_pt

        photo_cell = ws.cell(row=r, column=1, value="")
        photo_cell.alignment = center
        photo_cell.border = border

        article_cell = ws.cell(row=r, column=2, value=row.get("article") or "")
        article_cell.alignment = left_wrap
        article_cell.border = border

        sticker_cell = ws.cell(row=r, column=3, value=row.get("sticker") or "")
        sticker_cell.alignment = center
        sticker_cell.border = border

        barcode_cell = ws.cell(row=r, column=4, value=row.get("barcode") or "")
        barcode_cell.alignment = center
        barcode_cell.border = border

        box_cell = ws.cell(row=r, column=5, value="")
        box_cell.alignment = center
        box_cell.border = border

        photo_bytes = _fetch_product_photo_bytes(row.get("photo"))
        if photo_bytes:
            try:
                from PIL import Image  # type: ignore

                img = Image.open(io.BytesIO(photo_bytes)).convert("RGB")
                img.thumbnail((photo_width_px, photo_height_px), Image.Resampling.LANCZOS)
                canvas = Image.new("RGB", (photo_width_px, photo_height_px), "white")
                offset_x = (photo_width_px - img.width) // 2
                offset_y = (photo_height_px - img.height) // 2
                canvas.paste(img, (offset_x, offset_y))
                buf = io.BytesIO()
                canvas.save(buf, format="PNG")
                buf.seek(0)
                xl_img = XLImage(buf)
                xl_img.width = photo_width_px
                xl_img.height = photo_height_px
                ws.add_image(xl_img, f"A{r}")
            except Exception:
                pass

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _load_supply_print_context(token: str, supply_id: str, user_id: int) -> dict[str, Any]:
    """Загружает данные для страницы печати / экспорта PDF поставки FBS."""
    raw_items, stickers_err = _fetch_supply_orders_raw(token, supply_id)
    items = _normalize_supply_order_items(raw_items)
    order_ids = [it.get("id") for it in items if it.get("id")]
    meta_by_id: dict[int, dict[str, Any]] = {}
    for it in items:
        try:
            oid = int(it.get("id"))
            meta_by_id[oid] = it
        except Exception:
            continue

    stickers_raw, stickers_fetch_err = _fetch_fbs_order_stickers(token, order_ids)
    stickers: list[dict[str, Any]] = []
    for s in stickers_raw:
        oid = s.get("orderId")
        meta = {}
        try:
            meta = meta_by_id.get(int(oid), {}) if oid is not None else {}
        except Exception:
            meta = {}
        stickers.append(
            {
                "order_id": oid,
                "article": meta.get("article") or "",
                "barcode": meta.get("barcode") or s.get("barcode") or "",
                "part_a": s.get("partA") or "",
                "part_b": s.get("partB") or "",
                "file": s.get("file") or "",
            }
        )

    barcode = _resolve_supply_barcode(token, supply_id)
    trbx_items = _load_trbx_items(token, supply_id)
    product_labels = _build_product_barcode_labels(raw_items, items, token, user_id)
    order_tape_pairs = _build_order_tape_pairs(product_labels, stickers)

    return {
        "raw_items": raw_items,
        "items": items,
        "stickers": stickers,
        "stickers_error": stickers_fetch_err or stickers_err,
        "barcode": barcode,
        "trbx_items": trbx_items,
        "product_labels": product_labels,
        "order_tape_pairs": order_tape_pairs,
        "items_count": len(items),
    }


@fbs_supplies_bp.route("/api/fbs/supplies", methods=["GET"])
@login_required
def api_fbs_supplies():
    """Список поставок FBS (с пагинацией и возможностью обновления)."""
    token = effective_wb_api_token(current_user)
    if not token:
        return jsonify({"items": [], "lastUpdated": None}), 200

    refresh_flag = request.args.get("refresh") in ("1", "true", "True")
    limit_param = request.args.get("limit", default="5")
    offset_param = request.args.get("offset", default="0")
    try:
        limit_i = max(1, min(1000, int(limit_param)))
    except Exception:
        limit_i = 5
    try:
        offset_i = int(offset_param)
    except Exception:
        offset_i = 0

    # Всегда пробуем загрузить из API с fallback на кэш
    all_supplies_raw: List[Dict[str, Any]] = []
    try:
        headers_list = [
            {"Authorization": f"{token}"},
            {"Authorization": f"Bearer {token}"},
        ]
        for hdrs in headers_list:
            try:
                resp = get_with_retry(
                    FBS_SUPPLIES_LIST_URL,
                    hdrs,
                    params={"limit": 1000, "next": 0},
                    timeout_s=10,
                )
                data = resp.json()
                print(f"FBS supplies API response: type={type(data)}, keys={list(data.keys()) if isinstance(data, dict) else 'not dict'}")
                if isinstance(data, list):
                    all_supplies_raw = data
                    print(f"Got {len(all_supplies_raw)} supplies from list response")
                elif isinstance(data, dict):
                    all_supplies_raw = (
                        data.get("supplies", []) or data.get("data", []) or []
                    )
                    print(f"Got {len(all_supplies_raw)} supplies from dict response")
                    if all_supplies_raw and isinstance(all_supplies_raw[0], dict):
                        print(f"First supply sample keys: {list(all_supplies_raw[0].keys())}")
                        print(f"First supply sample: {all_supplies_raw[0]}")
                else:
                    continue
                break
            except requests.RequestException:
                continue
        if not all_supplies_raw:
            cached = load_fbs_supplies_cache() or {}
            all_supplies_raw = cached.get("all_supplies_raw", [])
    except Exception:
        cached = load_fbs_supplies_cache() or {}
        all_supplies_raw = cached.get("all_supplies_raw", [])

    # Сортируем по дате создания (новые сверху)
    try:
        all_supplies_raw.sort(key=lambda x: x.get("createdAt", ""), reverse=True)
    except Exception:
        pass

    # Получаем все заказы и считаем количество для каждой поставки
    supply_counts: Dict[str, int] = {}
    try:
        headers_list = [
            {"Authorization": f"{token}"},
            {"Authorization": f"Bearer {token}"},
        ]
        for hdrs in headers_list:
            try:
                orders_url = FBS_ORDERS_URL
                orders_params = {"limit": 1000, "next": 0}
                orders_resp = requests.get(orders_url, headers=hdrs, params=orders_params, timeout=30)
                if orders_resp.status_code == 200:
                    orders_data = orders_resp.json()
                    all_orders: List[Dict[str, Any]] = []
                    if isinstance(orders_data, dict):
                        if isinstance(orders_data.get("orders"), list):
                            all_orders = orders_data["orders"]
                    elif isinstance(orders_data, list):
                        all_orders = [it for it in orders_data if isinstance(it, dict)]
                    
                    # Группируем заказы по supplyId
                    for order in all_orders:
                        if not isinstance(order, dict):
                            continue
                        order_supply_id = None
                        for field in ["supplyId", "supply_id", "supplyID", "supply"]:
                            if field in order:
                                order_supply_id = str(order[field])
                                break
                        if order_supply_id:
                            supply_counts[order_supply_id] = supply_counts.get(order_supply_id, 0) + 1
                    break
            except Exception:
                continue
    except Exception:
        pass

    supplies_to_process = all_supplies_raw[offset_i : offset_i + limit_i]

    # Нормализуем для фронтенда
    norm_items: List[Dict[str, Any]] = []
    for it in supplies_to_process:
        if not isinstance(it, dict):
            continue
        supply_id = it.get("id") or it.get("supplyId") or it.get("supply_id")
        if not supply_id:
            continue

        # Количество товаров (из API или из подсчета заказов)
        count = it.get("orderCount") or it.get("ordersCount") or it.get("count")
        if count is None or count == 0:
            # Используем подсчитанное количество из заказов
            count = supply_counts.get(supply_id, 0)

        # Информация о датах и статусе
        created_raw = it.get("createdAt") or it.get("dateCreated") or it.get("date")
        closed_at = it.get("closedAt") or it.get("doneAt")
        raw_status = str(it.get("status") or "").upper()
        done_flag = bool(it.get("done")) or raw_status in ("DONE", "CLOSED", "COMPLETED", "FINISHED", "SHIPPED")

        # Форматируем даты
        def _fmt(raw):
            if not raw:
                return ""
            try:
                dt = parse_wb_datetime(str(raw))
                dt_msk = to_moscow(dt) if dt else None
                return (
                    dt_msk.strftime("%d.%m.%Y %H:%M")
                    if dt_msk
                    else (str(raw) if raw else "")
                )
            except Exception:
                return str(raw)

        created_str = _fmt(created_raw)
        
        # Status label for UI
        if done_flag:
            status_label = "Отгружено"
            try:
                status_dt_str = _fmt(closed_at) if closed_at else ""
            except Exception:
                status_dt_str = str(closed_at) if closed_at else ""
        else:
            status_label = "Не отгружена"
            status_dt_str = ""

        norm_items.append(
            {
                "supplyId": supply_id,
                "date": created_str,
                "count": count,
                "status": status_label,
                "statusDt": status_dt_str,
            }
        )

    # Сохраняем кэш
    try:
        save_fbs_supplies_cache({"all_supplies_raw": all_supplies_raw})
    except Exception:
        pass

    last_updated = (
        datetime.now(MOSCOW_TZ).strftime("%d.%m.%Y %H:%M") if norm_items else None
    )

    return jsonify(
        {
            "items": norm_items,
            "total": len(all_supplies_raw),
            "hasMore": offset_i + limit_i < len(all_supplies_raw),
            "lastUpdated": last_updated,
        }
    )


def _fetch_supply_orders_raw(token: str, supply_id: str) -> tuple[List[Dict[str, Any]], str | None]:
    """Загружает сырые заказы поставки FBS через /api/v3/orders."""
    headers_list = [
        {"Authorization": f"{token}"},
        {"Authorization": f"Bearer {token}"},
    ]
    last_err: str | None = None
    items: List[Dict[str, Any]] = []
    supply_id_fields = ["supplyId", "supply_id", "supplyID", "supply"]

    for idx, hdrs in enumerate(headers_list):
        try:
            orders_params = {"limit": 1000, "next": 0}
            orders_resp = requests.get(
                FBS_ORDERS_URL, headers=hdrs, params=orders_params, timeout=30
            )
            if orders_resp.status_code != 200:
                continue

            orders_data = orders_resp.json()
            all_orders: List[Dict[str, Any]] = []
            if isinstance(orders_data, dict):
                if isinstance(orders_data.get("orders"), list):
                    all_orders = orders_data["orders"]
                elif isinstance(orders_data.get("data"), list):
                    all_orders = orders_data["data"]
            elif isinstance(orders_data, list):
                all_orders = [it for it in orders_data if isinstance(it, dict)]

            filtered_items: List[Dict[str, Any]] = []
            for order in all_orders:
                if not isinstance(order, dict):
                    continue
                order_supply_id = None
                for field in supply_id_fields:
                    if field in order:
                        order_supply_id = order[field]
                        break
                if order_supply_id and str(order_supply_id) == str(supply_id):
                    filtered_items.append(order)

            if filtered_items:
                return filtered_items, None
            items = filtered_items
        except Exception as e:
            last_err = str(e)
            continue

    return items, last_err


def _normalize_supply_order_items(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Нормализует заказы поставки и обогащает из кэша товаров."""
    norm: List[Dict[str, Any]] = []
    prod_cached = load_products_cache() or {}
    by_nm: Dict[int, Dict[str, Any]] = {}
    try:
        for it in (prod_cached.get("items") or []):
            nmv = it.get("nm_id") or it.get("nmID")
            if nmv:
                by_nm[int(nmv)] = it
    except Exception:
        pass

    for it in items:
        if not isinstance(it, dict):
            continue
        nm = it.get("nmId") or it.get("nmID")
        photo = None
        barcode = None
        created_raw = it.get("createdAt") or it.get("dateCreated") or it.get("date")
        try:
            _dt = parse_wb_datetime(str(created_raw)) if created_raw else None
            _dt_msk = to_moscow(_dt) if _dt else None
            created_str = (
                _dt_msk.strftime("%d.%m.%Y %H:%M")
                if _dt_msk
                else (str(created_raw) if created_raw else "")
            )
        except Exception:
            created_str = str(created_raw) if created_raw else ""

        if nm:
            try:
                hit = by_nm.get(int(nm))
            except Exception:
                hit = None
            if hit:
                photo = hit.get("photo")
                if hit.get("barcode"):
                    barcode = hit.get("barcode")
                elif isinstance(hit.get("barcodes"), list) and hit.get("barcodes"):
                    barcode = str(hit.get("barcodes")[0])
                else:
                    sizes = hit.get("sizes") or []
                    if isinstance(sizes, list):
                        for s in sizes:
                            bar_list = s.get("skus") or s.get("barcodes")
                            if isinstance(bar_list, list) and bar_list:
                                barcode = str(bar_list[0])
                                break

        norm.append(
            {
                "id": it.get("id") or it.get("orderId"),
                "article": it.get("article") or it.get("supplierArticle"),
                "barcode": barcode or it.get("barcode") or "",
                "nm_id": nm,
                "photo": photo,
                "createdAt": created_str,
            }
        )
    return norm


def _group_supply_items_for_export(items: List[Dict[str, Any]]) -> List[tuple[str, str, str, int]]:
    """Группирует товары поставки для экспорта: (наименование, nm_id, баркод, кол-во)."""
    agg: Dict[tuple[str, str, str], int] = {}
    for it in items:
        name = str(it.get("article") or "").strip()
        nm_id = str(it.get("nm_id") or "").strip()
        barcode = str(it.get("barcode") or "").strip()
        key = (name, nm_id, barcode)
        agg[key] = agg.get(key, 0) + 1
    return [
        (name, nm_id, barcode, qty)
        for (name, nm_id, barcode), qty in sorted(
            agg.items(), key=lambda x: (-x[1], x[0][0])
        )
    ]


def _build_supply_xls(rows: List[tuple[str, str, str, int]], sheet_name: str = "Поставка") -> io.BytesIO:
    """Собирает .xls с группированными товарами."""
    try:
        import xlwt  # type: ignore
    except Exception as exc:
        raise RuntimeError("На сервере отсутствует зависимость xlwt (для .xls)") from exc

    wb = xlwt.Workbook()
    ws = wb.add_sheet(sheet_name[:31])
    header_style = xlwt.easyxf("font: bold on; align: horiz center")
    num_style = xlwt.easyxf("align: horiz right")
    ws.write(0, 0, "Наименование", header_style)
    ws.write(0, 1, "Артикул WB (nmId)", header_style)
    ws.write(0, 2, "Баркод", header_style)
    ws.write(0, 3, "Количество", header_style)
    for row_idx, (name, nm_id, barcode, qty) in enumerate(rows, start=1):
        ws.write(row_idx, 0, name)
        ws.write(row_idx, 1, nm_id)
        ws.write(row_idx, 2, barcode)
        ws.write(row_idx, 3, qty, num_style)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


@fbs_supplies_bp.route("/api/fbs/supplies/<supply_id>/orders", methods=["GET"])
@login_required
def api_fbs_supply_orders(supply_id: str):
    """Состав (товары) конкретной поставки FBS."""
    token = effective_wb_api_token(current_user)
    if not token:
        return jsonify({"items": []}), 200

    try:
        items, last_err = _fetch_supply_orders_raw(token, supply_id)
        norm = _normalize_supply_order_items(items)
        return jsonify({"items": norm}), 200
    except Exception as e:
        return jsonify({"items": [], "error": str(e)}), 200


@fbs_supplies_bp.route("/api/fbs/supplies/<supply_id>/export", methods=["GET"])
@login_required
def api_fbs_supply_export(supply_id: str):
    """Экспорт состава поставки FBS в Excel (сгруппировано по товарам)."""
    token = effective_wb_api_token(current_user)
    if not token:
        return ("Требуется API токен", 400)

    try:
        items, _ = _fetch_supply_orders_raw(token, supply_id)
        norm = _normalize_supply_order_items(items)
        if not norm:
            return ("Нет товаров в поставке", 400)

        grouped = _group_supply_items_for_export(norm)
        out = _build_supply_xls(grouped, sheet_name=f"Поставка {supply_id}")
        safe_id = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in str(supply_id))
        filename = f"fbs_supply_{safe_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xls"
        return send_file(
            out,
            mimetype="application/vnd.ms-excel",
            as_attachment=True,
            download_name=filename,
        )
    except RuntimeError as exc:
        return (str(exc), 500)
    except Exception as exc:
        return (f"Ошибка: {exc}", 500)


@fbs_supplies_bp.route("/fbs/supplies/<supply_id>/print", methods=["GET"])
@login_required
def fbs_supply_print_page(supply_id: str):
    """Страница печати QR поставки и стикеров сборочных заданий."""
    token = effective_wb_api_token(current_user)
    if not token:
        return "Требуется API токен Wildberries", 400

    ctx = _load_supply_print_context(token, supply_id, current_user.id)

    return render_template(
        "fbs_supply_print.html",
        supply_id=supply_id,
        stickers=ctx["stickers"],
        stickers_error=ctx["stickers_error"],
        barcode=ctx["barcode"],
        trbx_items=ctx["trbx_items"],
        product_labels=ctx["product_labels"],
        order_tape_pairs=ctx["order_tape_pairs"],
        items_count=ctx["items_count"],
        autoprint=request.args.get("autoprint") in ("1", "true", "True"),
    )


@fbs_supplies_bp.route("/fbs/supplies/<supply_id>/pick-list", methods=["GET"])
@login_required
def fbs_supply_pick_list(supply_id: str):
    """Скачивание листа подбора поставки FBS в Excel."""
    token = effective_wb_api_token(current_user)
    if not token:
        return "Требуется API токен Wildberries", 400

    ctx = _load_supply_print_context(token, supply_id, current_user.id)
    rows = _build_pick_list_rows(ctx)
    if not rows:
        return ("Нет товаров в поставке", 400)

    try:
        xlsx_buf = _build_pick_list_xlsx(supply_id, rows)
    except Exception as exc:
        return (f"Ошибка создания Excel: {exc}", 500)

    filename = _dated_download_name("Лист подбора", "xlsx")
    return send_file(
        xlsx_buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@fbs_supplies_bp.route("/fbs/supplies/<supply_id>/pdf/print", methods=["GET"])
@login_required
def fbs_supply_print_pdf(supply_id: str):
    """PDF для печати из браузера — тот же рендер, что и при сохранении в PDF."""
    token = effective_wb_api_token(current_user)
    if not token:
        return "Требуется API токен Wildberries", 400

    qr = _truthy_arg(request.args.get("qr"))
    trbx = _truthy_arg(request.args.get("trbx"))
    stickers = _truthy_arg(request.args.get("stickers"))
    barcodes = _truthy_arg(request.args.get("barcodes"))

    if not (qr or trbx or stickers or barcodes):
        return ("Выберите хотя бы один тип стикеров для печати", 400)

    ctx = _load_supply_print_context(token, supply_id, current_user.id)
    try:
        pdf_buf = _build_fbs_print_pdf(
            ctx,
            qr=qr,
            trbx=trbx,
            stickers=stickers,
            barcodes=barcodes,
        )
    except ValueError as exc:
        return (str(exc), 400)
    except Exception as exc:
        return (f"Ошибка создания PDF: {exc}", 500)

    safe_id = _safe_supply_filename_part(supply_id)
    filename = f"fbs_supply_{safe_id}_print_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(
        pdf_buf,
        mimetype="application/pdf",
        as_attachment=False,
        download_name=filename,
    )


@fbs_supplies_bp.route("/fbs/supplies/<supply_id>/pdf/<block>", methods=["GET"])
@login_required
def fbs_supply_pdf(supply_id: str, block: str):
    """Скачивание PDF для блока печати: qr, trbx, tape."""
    token = effective_wb_api_token(current_user)
    if not token:
        return "Требуется API токен Wildberries", 400

    block_key = str(block or "").strip().lower()
    ctx = _load_supply_print_context(token, supply_id, current_user.id)
    images_b64: list[str] = []
    suffix = block_key

    if block_key == "qr":
        barcode = ctx.get("barcode") or {}
        if barcode.get("file"):
            images_b64.append(barcode["file"])
    elif block_key == "trbx":
        for item in ctx.get("trbx_items") or []:
            if isinstance(item, dict) and item.get("file"):
                images_b64.append(item["file"])
    elif block_key == "tape":
        pairs = ctx.get("order_tape_pairs") or []
        if not pairs:
            return ("Нет стикеров для сохранения в PDF", 400)
        try:
            pdf_buf = _build_tape_pdf(pairs)
        except Exception as exc:
            return (f"Ошибка создания PDF: {exc}", 500)
        filename = _dated_download_name("Стикеры и ШК товаров", "pdf")
        return send_file(
            pdf_buf,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename,
        )
    else:
        return ("Неизвестный тип PDF", 400)

    if not images_b64:
        return ("Нет стикеров для сохранения в PDF", 400)

    try:
        pdf_buf = _build_images_pdf(images_b64)
    except ValueError as exc:
        return (str(exc), 400)
    except Exception as exc:
        return (f"Ошибка создания PDF: {exc}", 500)

    if block_key == "trbx":
        filename = _dated_download_name("QR грузоместа", "pdf")
    elif block_key == "qr":
        filename = _dated_download_name("QR поставки", "pdf")
    else:
        filename = _dated_download_name(f"fbs_supply_{_safe_supply_filename_part(supply_id)}_{suffix}", "pdf")
    return send_file(
        pdf_buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


@fbs_supplies_bp.route("/api/fbs/supplies/<supply_id>/trbx", methods=["GET", "POST"])
@login_required
def api_fbs_supply_trbx(supply_id: str):
    """Список грузомест / добавление грузоместа в поставку FBS."""
    token = effective_wb_api_token(current_user)
    if not token:
        return jsonify({"error": "no_token"}), 401

    if request.method == "GET":
        items = _load_trbx_items(token, supply_id)
        return jsonify({"items": items, "count": len(items)}), 200

    payload = request.get_json(silent=True) or {}
    try:
        amount = int(payload.get("amount") or 1)
    except Exception:
        amount = 1

    new_ids, err = _add_fbs_supply_trbx(token, supply_id, amount=amount)
    if err and not new_ids:
        return jsonify({"error": err}), 502

    items = _fetch_fbs_trbx_stickers(token, supply_id, new_ids) if new_ids else []
    return jsonify({"success": True, "items": items, "trbxIds": new_ids}), 200


@fbs_supplies_bp.route("/api/fbs/supplies/<supply_id>/deliver", methods=["PATCH", "POST"])
@login_required
def api_fbs_supply_deliver(supply_id: str):
    """Передать поставку FBS в доставку (отгрузка на склад WB по QR поставки)."""
    token = effective_wb_api_token(current_user)
    if not token:
        return jsonify({"error": "no_token"}), 401

    url = FBS_SUPPLY_DELIVER_URL.replace("{supplyId}", str(supply_id))
    last_err: str | None = None
    for hdrs in _wb_auth_headers(token):
        try:
            resp = requests.patch(url, headers=hdrs, timeout=30)
            if resp.status_code in (200, 204):
                return jsonify({"success": True, "supplyId": supply_id}), 200
            last_err = f"HTTP {resp.status_code}: {resp.text[:300]}"
        except Exception as exc:
            last_err = str(exc)

    return jsonify({"error": last_err or "Unknown error"}), 502


@fbs_supplies_bp.route("/api/fbs/supplies/<supply_id>/orders/<order_id>", methods=["PATCH", "POST"])
@login_required
def api_fbs_add_order_to_supply(supply_id: str, order_id: str):
    """Добавить сборочное задание в поставку.

    Использует новый метод WB:
    PATCH /api/marketplace/v3/supplies/{supplyId}/orders
    c телом { "orders": [orderId] }.
    """
    import logging

    logger = logging.getLogger(__name__)
    logger.info(f"=== Adding order {order_id} to supply {supply_id} ===")

    token = effective_wb_api_token(current_user)
    if not token:
        logger.warning("No token provided")
        return jsonify({"error": "No token"}), 401

    headers_list = [
        {"Authorization": f"{token}"},
        {"Authorization": f"Bearer {token}"},
    ]

    url = FBS_SUPPLY_ADD_ORDERS_URL.replace("{supplyId}", str(supply_id))
    payload = {"orders": [int(order_id)]}
    last_err: str | None = None

    for hdrs in headers_list:
        try:
            hdrs_with_content = dict(hdrs)
            hdrs_with_content["Content-Type"] = "application/json"
            logger.info(f"PATCH {url} with payload={payload}")
            # Уменьшаем таймаут до 10 секунд, чтобы UI не «висел» по 30 секунд
            resp = requests.patch(url, headers=hdrs_with_content, json=payload, timeout=10)
            logger.info(f"WB response status={resp.status_code}, body={resp.text[:300]}")

            if resp.status_code in (200, 201, 204):
                return jsonify({"success": True}), 200
            if resp.status_code == 409:
                # Задание уже привязано к поставке
                return jsonify({"error": "Order already in supply"}), 409

            # Прочие ошибки — сохраняем текст и пробуем со следующим вариантом заголовков
            last_err = f"HTTP {resp.status_code}: {resp.text[:300]}"
        except Exception as e:
            import traceback

            last_err = str(e)
            logger.error(f"Exception while calling WB add-orders endpoint: {e}")
            traceback.print_exc()

    return jsonify({"error": last_err or "Unknown error"}), 500


