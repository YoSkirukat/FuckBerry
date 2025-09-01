import sys
import json
from pathlib import Path
from openpyxl import load_workbook


def cell_value(ws, r, c):
    try:
        return ws.cell(row=r, column=c).value
    except Exception:
        return None


def main():
    if len(sys.argv) > 1:
        path = Path(sys.argv[1])
    else:
        path = Path(__file__).resolve().parents[1] / "Расшифровка WB_с формулами.xlsx"
    result = {
        "path": str(path),
        "exists": path.exists(),
        "dashboard": {
            "rows": []
        }
    }
    if not path.exists():
        print(json.dumps(result, ensure_ascii=False))
        return
    # Load twice: formulas and cached values
    wb_formula = load_workbook(filename=str(path), data_only=False)
    wb_values = load_workbook(filename=str(path), data_only=True)
    if "DASHBOARD" not in wb_formula.sheetnames:
        if "Dashboard" in wb_formula.sheetnames:
            dash_name = "Dashboard"
        else:
            dash_name = wb_formula.sheetnames[0]
    else:
        dash_name = "DASHBOARD"
    ws_f = wb_formula[dash_name]
    ws_v = wb_values[dash_name]

    # Heuristic: read first 200 rows, 10 columns. Expect labels at col A, values at col B (can be extended later).
    rows_out = []
    for r in range(1, 201):
        label = cell_value(ws_f, r, 1)
        if label is None or (isinstance(label, str) and label.strip() == ""):
            continue
        val_formula = cell_value(ws_f, r, 2)
        val_value = cell_value(ws_v, r, 2)
        rows_out.append({
            "row": r,
            "label": str(label),
            "value": val_value,
            "formula": str(val_formula) if isinstance(val_formula, str) and val_formula.startswith('=') else None,
            "address": f"A{r}:B{r}"
        })

    result["dashboard"]["rows"] = rows_out
    print(json.dumps(result, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()


